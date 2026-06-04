import os
import torch
from openpyxl import Workbook, load_workbook


def save_checkpoint(state, checkpoint_dir, filename):
    """Save a training checkpoint to *checkpoint_dir/filename*."""
    os.makedirs(checkpoint_dir, exist_ok=True)
    path = os.path.join(checkpoint_dir, filename)
    torch.save(state, path)
    return path


def load_checkpoint(path, model, optimizer=None):
    """Load a checkpoint and restore model (and optionally optimizer) state.

    Returns:
        dict: The full checkpoint dictionary (contains 'epoch', 'val_loss', etc.)
    """
    checkpoint = torch.load(path, map_location="cpu")
    model.load_state_dict(checkpoint["model_state_dict"])
    if optimizer is not None and "optimizer_state_dict" in checkpoint:
        optimizer.load_state_dict(checkpoint["optimizer_state_dict"])
    return checkpoint


def save_epoch_metrics_to_excel(rows, output_path, headers=None):
    """Save epoch metrics rows to an Excel file."""
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "epoch_metrics"
    if headers is None:
        headers = list(rows[0].keys()) if rows else []
    if headers:
        worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    workbook.save(output_path)


def load_epoch_metrics_from_excel(path):
    """Load epoch metrics rows from an Excel file."""
    if not os.path.isfile(path):
        return [], []
    workbook = load_workbook(path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0]) if rows[0] else []
    data_rows = []
    for row in rows[1:]:
        if row is None:
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            row_dict[header] = row[idx] if idx < len(row) else None
        if any(value is not None for value in row_dict.values()):
            data_rows.append(row_dict)
    return data_rows, headers
